"""Excel workbook of target object state.

Uses openpyxl if available; otherwise falls back to a tiny, dependency-free
native .xlsx writer (an .xlsx is just a zip of XML parts). Either way the same
sheets are produced:
    Summary | Objects | Column Mapping | Effort
"""
from __future__ import annotations

from typing import List, Sequence

from ..analyzer import AnalysisResult


def write_excel(analysis: AnalysisResult, path: str) -> str:
    sheets = _build_sheets(analysis)
    try:
        import openpyxl  # noqa: F401
        _write_openpyxl(sheets, path)
    except ImportError:
        _write_native(sheets, path)
    return path


def _build_sheets(analysis: AnalysisResult):
    sc = analysis.status_counts()
    summary = [
        ["Metric", "Value"],
        ["Source engine", analysis.source_engine],
        ["Target engine", analysis.target_engine],
        ["Objects total", len(analysis.objects)],
        ["Average compatibility %", analysis.avg_compatibility()],
        ["Automatic (>=90%)", sc.get("Automatic", 0)],
        ["Review (60-89%)", sc.get("Review", 0)],
        ["Manual (<60%)", sc.get("Manual", 0)],
        ["Total effort (hours)", analysis.total_effort()],
        ["Total effort (person-days)", round(analysis.total_effort() / 8.0, 1)],
    ]

    objects = [["Object Type", "Name", "Compatibility %", "Target State",
                "Risk", "Effort (h)", "Detail"]]
    for o in analysis.objects:
        objects.append([o.object_type, o.name, o.compatibility, o.status,
                        o.risk, o.effort_hours, o.detail])

    columns = [["Table", "Column", "Source Type", "Target Type",
                "Compatibility %", "Nullable", "Note"]]
    for cm in analysis.column_mappings:
        columns.append([cm.table, cm.column, cm.source_type, cm.target_type,
                        cm.compatibility, "Y" if cm.nullable else "N", cm.note])

    effort = [["Object Type", "Effort (hours)"]]
    for k, v in analysis.effort_by_type().items():
        effort.append([k, v])
    effort.append(["TOTAL", analysis.total_effort()])

    return {
        "Summary": summary,
        "Objects": objects,
        "Column Mapping": columns,
        "Effort": effort,
    }


def _write_openpyxl(sheets: dict, path: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")

    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(title=sheet_name[:31])
        for r_idx, row in enumerate(rows, start=1):
            for c_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                if r_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font
        # autosize-ish
        for c_idx in range(1, len(rows[0]) + 1):
            width = max((len(str(row[c_idx - 1])) for row in rows if c_idx - 1 < len(row)),
                        default=10)
            ws.column_dimensions[get_column_letter(c_idx)].width = min(max(width + 2, 10), 60)
        ws.freeze_panes = "A2"
    wb.save(path)


# --------------------------------------------------------------------------- #
# Native fallback writer - minimal but valid .xlsx (zip of XML).
# --------------------------------------------------------------------------- #
def _write_native(sheets: dict, path: str) -> None:
    import zipfile
    from xml.sax.saxutils import escape

    def col_letter(n: int) -> str:
        s = ""
        while n > 0:
            n, r = divmod(n - 1, 26)
            s = chr(65 + r) + s
        return s

    def sheet_xml(rows: List[Sequence]) -> str:
        out = ['<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
               '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">',
               '<sheetData>']
        for r_idx, row in enumerate(rows, start=1):
            out.append(f'<row r="{r_idx}">')
            for c_idx, val in enumerate(row, start=1):
                ref = f"{col_letter(c_idx)}{r_idx}"
                if isinstance(val, (int, float)) and not isinstance(val, bool):
                    out.append(f'<c r="{ref}"><v>{val}</v></c>')
                else:
                    out.append(f'<c r="{ref}" t="inlineStr"><is><t>'
                               f'{escape(str(val))}</t></is></c>')
            out.append('</row>')
        out.append('</sheetData></worksheet>')
        return "".join(out)

    names = list(sheets.keys())
    content_types = ['<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
                     '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">',
                     '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>',
                     '<Default Extension="xml" ContentType="application/xml"/>',
                     '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>']
    for i in range(len(names)):
        content_types.append(
            f'<Override PartName="/xl/worksheets/sheet{i+1}.xml" '
            f'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>')
    content_types.append('</Types>')

    root_rels = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                 '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                 '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>'
                 '</Relationships>')

    wb_sheets = "".join(
        f'<sheet name="{escape(n)[:31]}" sheetId="{i+1}" r:id="rId{i+1}"/>'
        for i, n in enumerate(names))
    workbook = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
                'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
                f'<sheets>{wb_sheets}</sheets></workbook>')

    wb_rels = ['<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
               '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">']
    for i in range(len(names)):
        wb_rels.append(
            f'<Relationship Id="rId{i+1}" '
            f'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" '
            f'Target="worksheets/sheet{i+1}.xml"/>')
    wb_rels.append('</Relationships>')

    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("[Content_Types].xml", "".join(content_types))
        z.writestr("_rels/.rels", root_rels)
        z.writestr("xl/workbook.xml", workbook)
        z.writestr("xl/_rels/workbook.xml.rels", "".join(wb_rels))
        for i, n in enumerate(names):
            z.writestr(f"xl/worksheets/sheet{i+1}.xml", sheet_xml(sheets[n]))
